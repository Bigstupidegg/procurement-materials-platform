from __future__ import annotations

import hashlib
import io
import json
import math
import os
import re
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import openpyxl
import requests

ROOT = Path(__file__).resolve().parents[1]
MATERIALS_PATH = ROOT / "config" / "materials.json"
COLUMN_CONFIG_PATH = ROOT / "config" / "world-bank-columns.json"
WORLD_BANK_OUTPUT_PATH = ROOT / "data" / "world-bank.json"
STATUS_OUTPUT_PATH = ROOT / "data" / "status.json"

MIN_XLSX_BYTES = 50_000
MAX_XLSX_BYTES = 50 * 1024 * 1024
MIN_OBSERVATIONS_PER_SERIES = 60
PERIOD_PATTERN = re.compile(r"^(?P<year>\d{4})M(?P<month>0[1-9]|1[0-2])$")
MISSING_TEXT_VALUES = {"", "...", "…", "..", "-", "na", "n/a", "null"}


class MonthlyLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        for key, value in attrs:
            if key.lower() == "href" and value:
                self.hrefs.append(value)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temp_path, path)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any) -> str:
    return normalize_text(value).casefold()


def validate_source_url(url: str, allowed_hosts: set[str]) -> None:
    parsed = urlparse(url)
    if parsed.scheme != "https" or parsed.hostname not in allowed_hosts:
        raise RuntimeError(f"來源網址不在允許清單：{url}")


def discover_monthly_url(config: dict[str, Any], session: requests.Session) -> str:
    override_url = os.getenv("WORLD_BANK_MONTHLY_URL", "").strip()
    allowed_hosts = set(config["sourceUrlAllowlistDomains"])

    if override_url:
        validate_source_url(override_url, allowed_hosts)
        return override_url

    landing_url = config["landingPageUrl"]
    validate_source_url(landing_url, allowed_hosts | {"www.worldbank.org", "worldbank.org"})

    try:
        response = session.get(
            landing_url,
            headers={"User-Agent": "procurement-materials-platform/1.0 (+GitHub Actions)"},
            timeout=(15, 45),
        )
        response.raise_for_status()
        parser = MonthlyLinkParser()
        parser.feed(response.text)

        candidates: list[str] = []
        for href in parser.hrefs:
            absolute = urljoin(response.url, href)
            parsed = urlparse(absolute)
            filename = Path(parsed.path).name.casefold()
            if filename == "cmo-historical-data-monthly.xlsx":
                candidates.append(absolute)

        for candidate in candidates:
            try:
                validate_source_url(candidate, allowed_hosts)
                return candidate
            except RuntimeError:
                continue
    except requests.RequestException as exc:
        print(f"Warning: 無法從World Bank入口頁自動取得Excel連結，改用備援網址：{exc}")

    fallback_url = config["fallbackMonthlyWorkbookUrl"]
    validate_source_url(fallback_url, allowed_hosts)
    return fallback_url


def download_xlsx(url: str, config: dict[str, Any], session: requests.Session) -> tuple[bytes, dict[str, Any]]:
    allowed_hosts = set(config["sourceUrlAllowlistDomains"])
    validate_source_url(url, allowed_hosts)

    headers = {
        "User-Agent": "procurement-materials-platform/1.0 (+GitHub Actions)",
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream"
        ),
    }
    with session.get(url, headers=headers, stream=True, timeout=(15, 90), allow_redirects=True) as response:
        response.raise_for_status()
        final_url = response.url
        validate_source_url(final_url, allowed_hosts)

        content_type = response.headers.get("Content-Type", "").split(";", 1)[0].strip().lower()
        allowed_types = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        }
        if content_type and content_type not in allowed_types:
            raise RuntimeError(f"Content-Type異常：{content_type}")

        chunks: list[bytes] = []
        total = 0
        sha256 = hashlib.sha256()
        for chunk in response.iter_content(chunk_size=256 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > MAX_XLSX_BYTES:
                raise RuntimeError(f"檔案超過{MAX_XLSX_BYTES} bytes上限")
            sha256.update(chunk)
            chunks.append(chunk)

        raw = b"".join(chunks)
        metadata = {
            "requestedUrl": url,
            "finalUrl": final_url,
            "httpStatus": response.status_code,
            "contentType": content_type or None,
            "contentLengthHeader": response.headers.get("Content-Length"),
            "etag": response.headers.get("ETag"),
            "lastModified": response.headers.get("Last-Modified"),
            "fileSizeBytes": len(raw),
            "sha256": sha256.hexdigest(),
        }

    if len(raw) < MIN_XLSX_BYTES:
        raise RuntimeError(f"下載檔案過小：{len(raw)} bytes")
    if not raw.startswith(b"PK"):
        raise RuntimeError("下載內容不是有效的XLSX/ZIP檔案")
    return raw, metadata


def parse_period(value: Any) -> str | None:
    text = normalize_text(value)
    match = PERIOD_PATTERN.fullmatch(text)
    if not match:
        return None
    return f"{match.group('year')}-{match.group('month')}"


def to_finite_positive_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = normalize_text(value)
        if text.casefold() in MISSING_TEXT_VALUES:
            return None
        text = text.replace(",", "")
        try:
            value = float(text)
        except ValueError:
            return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isfinite(number) and number > 0:
            return number
    return None


def resolve_columns(
    header_values: list[Any],
    materials: list[dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, int]:
    normalized_headers: dict[str, list[int]] = {}
    for index, value in enumerate(header_values, start=1):
        key = normalize_header(value)
        if key:
            normalized_headers.setdefault(key, []).append(index)

    aliases = config.get("columnAliases", {})
    resolved: dict[str, int] = {}
    for material in materials:
        canonical = material["worldBankColumn"]
        candidates = [canonical, *aliases.get(canonical, [])]
        matches: list[int] = []
        for candidate in candidates:
            matches.extend(normalized_headers.get(normalize_header(candidate), []))
        unique_matches = sorted(set(matches))
        if not unique_matches:
            raise RuntimeError(f"缺少必要欄位：{canonical}")
        if len(unique_matches) > 1:
            raise RuntimeError(f"欄位名稱不唯一：{canonical}，候選欄位={unique_matches}")
        resolved[material["id"]] = unique_matches[0]
    return resolved


def parse_workbook(
    raw: bytes,
    download_metadata: dict[str, Any],
    materials: list[dict[str, Any]],
    config: dict[str, Any],
    generated_at: str,
) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(
        io.BytesIO(raw),
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    sheet_name = config["sheetName"]
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"找不到工作表：{sheet_name}")

    worksheet = workbook[sheet_name]
    header_row = int(config["headerRow"])
    unit_row = int(config["unitRow"])
    data_start_row = int(config["dataStartRow"])
    period_column = int(config["periodColumnIndex"])
    updated_text_row = int(config["updatedTextRow"])

    header_values = list(
        next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    )
    unit_values = list(
        next(worksheet.iter_rows(min_row=unit_row, max_row=unit_row, values_only=True))
    )
    column_indexes = resolve_columns(header_values, materials, config)

    workbook_updated_text = normalize_text(worksheet.cell(updated_text_row, 1).value)
    updated_match = re.search(r"Updated on\s+(.+)$", workbook_updated_text, flags=re.IGNORECASE)
    workbook_updated_on = None
    if updated_match:
        try:
            workbook_updated_on = datetime.strptime(updated_match.group(1), "%B %d, %Y").date().isoformat()
        except ValueError:
            workbook_updated_on = None

    observations: dict[str, list[dict[str, Any]]] = {material["id"]: [] for material in materials}
    recognized_periods: list[str] = []

    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        period = parse_period(row[period_column - 1] if len(row) >= period_column else None)
        if period is None:
            continue
        recognized_periods.append(period)

        for material in materials:
            column_index = column_indexes[material["id"]]
            raw_value = row[column_index - 1] if len(row) >= column_index else None
            value = to_finite_positive_float(raw_value)
            if value is None:
                continue
            observations[material["id"]].append(
                {
                    "period": period,
                    "value": int(value) if value.is_integer() else value,
                }
            )

    if len(recognized_periods) < MIN_OBSERVATIONS_PER_SERIES:
        raise RuntimeError(f"可辨識月份不足：{len(recognized_periods)}")
    if recognized_periods != sorted(recognized_periods):
        raise RuntimeError("月份順序不是遞增排列")
    if len(recognized_periods) != len(set(recognized_periods)):
        raise RuntimeError("月份資料有重複")

    latest_period = recognized_periods[-1]
    series_payload: dict[str, Any] = {}
    for material in materials:
        material_id = material["id"]
        series_observations = observations[material_id]
        if len(series_observations) < MIN_OBSERVATIONS_PER_SERIES:
            raise RuntimeError(
                f"{material['worldBankColumn']}有效資料不足：{len(series_observations)}"
            )
        if series_observations[-1]["period"] != latest_period:
            raise RuntimeError(
                f"{material['worldBankColumn']}最新月份缺值；"
                f"資料最新={series_observations[-1]['period']}，工作簿最新={latest_period}"
            )

        column_index = column_indexes[material_id]
        source_unit = normalize_text(unit_values[column_index - 1] if len(unit_values) >= column_index else "")
        if not source_unit:
            raise RuntimeError(f"{material['worldBankColumn']}缺少單位列")

        series_payload[material_id] = {
            "id": material_id,
            "nameZh": material["nameZh"],
            "nameEn": material["nameEn"],
            "currency": material["currency"],
            "displayUnit": material["unit"],
            "sourceColumn": material["worldBankColumn"],
            "sourceUnit": source_unit,
            "isLmeDerived": bool(material.get("isLmeDerived", False)),
            "attributionNote": material["attributionNote"],
            "observationCount": len(series_observations),
            "firstPeriod": series_observations[0]["period"],
            "latestPeriod": series_observations[-1]["period"],
            "observations": series_observations,
        }

    return {
        "schemaVersion": 1,
        "generatedAt": generated_at,
        "source": "WORLD_BANK_PINK_SHEET",
        "isRealData": True,
        "frequency": "monthly",
        "dataset": {
            "name": "World Bank Commodity Price Data (The Pink Sheet)",
            "sheetName": sheet_name,
            "currencyBasis": "nominal US dollars",
            "workbookUpdatedText": workbook_updated_text or None,
            "workbookUpdatedOn": workbook_updated_on,
            "latestPeriod": latest_period,
            "download": download_metadata,
        },
        "series": series_payload,
    }


def build_status(
    generated_at: str,
    world_bank_payload: dict[str, Any],
    previous_status: dict[str, Any] | None,
) -> dict[str, Any]:
    previous_status = previous_status or {}
    fred_status = previous_status.get(
        "fred",
        {
            "status": "NOT_CONFIGURED",
            "lastSuccessAt": None,
        },
    )
    dataset = world_bank_payload["dataset"]

    return {
        "schemaVersion": 1,
        "generatedAt": generated_at,
        "dataMode": "WORLD_BANK_PRIMARY",
        "worldBank": {
            "status": "SUCCESS",
            "lastAttemptAt": generated_at,
            "lastSuccessAt": generated_at,
            "latestPeriod": dataset["latestPeriod"],
            "sourceUpdatedOn": dataset["workbookUpdatedOn"],
            "sourceSha256": dataset["download"]["sha256"],
            "sourceUrl": dataset["download"]["finalUrl"],
        },
        "fred": fred_status,
        "isStale": False,
        "warnings": [],
    }


def main() -> None:
    generated_at = utc_now_iso()
    materials = read_json(MATERIALS_PATH)
    config = read_json(COLUMN_CONFIG_PATH)

    session = requests.Session()
    source_url = discover_monthly_url(config, session)
    print(f"World Bank monthly workbook: {source_url}")

    raw, download_metadata = download_xlsx(source_url, config, session)
    payload = parse_workbook(
        raw=raw,
        download_metadata=download_metadata,
        materials=materials,
        config=config,
        generated_at=generated_at,
    )

    previous_status = None
    if STATUS_OUTPUT_PATH.exists():
        try:
            previous_status = read_json(STATUS_OUTPUT_PATH)
        except (OSError, json.JSONDecodeError):
            previous_status = None

    status_payload = build_status(generated_at, payload, previous_status)

    # 所有下載、欄位解析與驗證成功後才原子置換，避免失敗覆蓋上一版有效資料。
    write_json_atomic(WORLD_BANK_OUTPUT_PATH, payload)
    write_json_atomic(STATUS_OUTPUT_PATH, status_payload)

    print(
        "World Bank sync success: "
        f"latestPeriod={payload['dataset']['latestPeriod']}, "
        f"series={len(payload['series'])}, "
        f"sha256={payload['dataset']['download']['sha256']}"
    )


if __name__ == "__main__":
    main()
