from __future__ import annotations

import hashlib
import io
import json
import math
import os
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests

WORLD_BANK_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)
ALLOWED_HOSTS = {"thedocs.worldbank.org", "pubdocs.worldbank.org"}
MAX_BYTES = 50 * 1024 * 1024
OUTPUT_DIR = Path("artifacts")


def normalize_cell(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return value
    if value is None or isinstance(value, (str, int, bool)):
        return value
    return str(value)


def download_xlsx(url: str) -> tuple[bytes, dict]:
    parsed = urlparse(url)
    if parsed.scheme != "https" or parsed.hostname not in ALLOWED_HOSTS:
        raise RuntimeError(f"來源網址不在允許清單：{url}")

    headers = {
        "User-Agent": "procurement-materials-platform/1.0 (+GitHub Actions)",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream",
    }
    with requests.get(url, headers=headers, stream=True, timeout=(15, 60), allow_redirects=True) as response:
        response.raise_for_status()
        final_url = response.url
        final_host = urlparse(final_url).hostname
        if final_host not in ALLOWED_HOSTS:
            raise RuntimeError(f"重新導向後網域不在允許清單：{final_host}")

        content_type = response.headers.get("Content-Type", "").split(";", 1)[0].strip().lower()
        allowed_types = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        }
        if content_type not in allowed_types:
            raise RuntimeError(f"Content-Type異常：{content_type or '(empty)'}")

        chunks: list[bytes] = []
        total = 0
        sha256 = hashlib.sha256()
        for chunk in response.iter_content(chunk_size=256 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > MAX_BYTES:
                raise RuntimeError(f"檔案超過{MAX_BYTES} bytes上限")
            sha256.update(chunk)
            chunks.append(chunk)

    raw = b"".join(chunks)
    if not raw.startswith(b"PK"):
        raise RuntimeError("下載內容不是有效的XLSX/ZIP檔案")

    metadata = {
        "requestedUrl": url,
        "finalUrl": final_url,
        "httpStatus": response.status_code,
        "contentType": content_type,
        "contentLengthHeader": response.headers.get("Content-Length"),
        "fileSizeBytes": len(raw),
        "sha256": sha256.hexdigest(),
        "etag": response.headers.get("ETag"),
        "lastModified": response.headers.get("Last-Modified"),
    }
    return raw, metadata


def non_empty_rows(ws, limit: int = 20):
    rows = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [normalize_cell(value) for value in row]
        if any(value not in (None, "") for value in values):
            rows.append({"row": row_number, "values": values})
            if len(rows) >= limit:
                break
    return rows


def tail_non_empty_rows(ws, limit: int = 8):
    collected = []
    start = max(1, ws.max_row - 100)
    for row_number, row in enumerate(
        ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True), start=start
    ):
        values = [normalize_cell(value) for value in row]
        if any(value not in (None, "") for value in values):
            collected.append({"row": row_number, "values": values})
    return collected[-limit:]


def find_candidate_header_rows(ws, keywords: set[str], max_scan_rows: int = 30):
    candidates = []
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True), start=1
    ):
        cells = [str(value).strip() for value in row if value not in (None, "")]
        lower_cells = [cell.lower() for cell in cells]
        matches = sorted({keyword for keyword in keywords if any(keyword in cell for cell in lower_cells)})
        if matches:
            candidates.append({"row": row_number, "matches": matches, "cells": cells})
    return candidates


def inspect_workbook(raw: bytes, metadata: dict) -> dict:
    workbook = openpyxl.load_workbook(
        io.BytesIO(raw), data_only=True, read_only=True, keep_links=False
    )
    keywords = {
        "copper",
        "zinc",
        "aluminum",
        "aluminium",
        "nickel",
        "iron ore",
        "crude oil",
        "brent",
        "natural gas",
        "monthly",
    }

    sheets = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheets.append(
            {
                "name": sheet_name,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "firstNonEmptyRows": non_empty_rows(ws),
                "lastNonEmptyRows": tail_non_empty_rows(ws),
                "candidateHeaderRows": find_candidate_header_rows(ws, keywords),
            }
        )

    return {
        "inspectedAtUtc": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "download": metadata,
        "sheetNames": workbook.sheetnames,
        "sheets": sheets,
    }


def make_text_report(report: dict) -> str:
    lines = [
        "World Bank Pink Sheet Monthly Workbook Inspection",
        "=" * 52,
        f"Requested URL: {report['download']['requestedUrl']}",
        f"Final URL: {report['download']['finalUrl']}",
        f"Content-Type: {report['download']['contentType']}",
        f"File size: {report['download']['fileSizeBytes']} bytes",
        f"SHA-256: {report['download']['sha256']}",
        "",
        "Sheet names:",
    ]
    for sheet in report["sheets"]:
        lines.append(f"- {sheet['name']} (rows={sheet['maxRow']}, columns={sheet['maxColumn']})")
        for candidate in sheet["candidateHeaderRows"]:
            lines.append(
                f"  Candidate row {candidate['row']} matches {', '.join(candidate['matches'])}: "
                + " | ".join(candidate["cells"][:30])
            )
    lines.append("")
    lines.append("完整儲存格內容請查看 world-bank-inspection.json。")
    return "\n".join(lines) + "\n"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    raw, metadata = download_xlsx(WORLD_BANK_URL)
    report = inspect_workbook(raw, metadata)

    json_path = OUTPUT_DIR / "world-bank-inspection.json"
    txt_path = OUTPUT_DIR / "world-bank-inspection.txt"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    txt_path.write_text(make_text_report(report), encoding="utf-8")

    print(txt_path.read_text(encoding="utf-8"))
    print(f"Generated: {json_path}")
    print(f"Generated: {txt_path}")


if __name__ == "__main__":
    main()
