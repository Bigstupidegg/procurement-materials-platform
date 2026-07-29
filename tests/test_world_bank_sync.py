from __future__ import annotations

import io
import unittest

import openpyxl

from scripts.sync_world_bank import parse_workbook, resolve_columns


MATERIALS = [
    {
        "id": "zinc",
        "nameZh": "鋅",
        "nameEn": "Zinc",
        "currency": "USD",
        "unit": "公噸(MT)",
        "worldBankColumn": "Zinc",
        "isLmeDerived": True,
        "attributionNote": "test",
    },
    {
        "id": "copper",
        "nameZh": "銅",
        "nameEn": "Copper",
        "currency": "USD",
        "unit": "公噸(MT)",
        "worldBankColumn": "Copper",
        "isLmeDerived": True,
        "attributionNote": "test",
    },
    {
        "id": "aluminium",
        "nameZh": "鋁",
        "nameEn": "Aluminium",
        "currency": "USD",
        "unit": "公噸(MT)",
        "worldBankColumn": "Aluminum",
        "isLmeDerived": True,
        "attributionNote": "test",
    },
    {
        "id": "nickel",
        "nameZh": "鎳",
        "nameEn": "Nickel",
        "currency": "USD",
        "unit": "公噸(MT)",
        "worldBankColumn": "Nickel",
        "isLmeDerived": True,
        "attributionNote": "test",
    },
    {
        "id": "iron_ore",
        "nameZh": "鐵礦砂",
        "nameEn": "Iron Ore",
        "currency": "USD",
        "unit": "乾公噸單位(dmtu)",
        "worldBankColumn": "Iron ore, cfr spot",
        "isLmeDerived": False,
        "attributionNote": "test",
    },
    {
        "id": "crude_oil",
        "nameZh": "原油",
        "nameEn": "Crude Oil, Brent",
        "currency": "USD",
        "unit": "桶(bbl)",
        "worldBankColumn": "Crude oil, Brent",
        "isLmeDerived": False,
        "attributionNote": "test",
    },
    {
        "id": "natural_gas",
        "nameZh": "天然氣",
        "nameEn": "Natural Gas, U.S.",
        "currency": "USD",
        "unit": "MMBtu",
        "worldBankColumn": "Natural gas, US",
        "isLmeDerived": False,
        "attributionNote": "test",
    },
]

CONFIG = {
    "sheetName": "Monthly Prices",
    "headerRow": 5,
    "unitRow": 6,
    "dataStartRow": 7,
    "periodColumnIndex": 1,
    "updatedTextRow": 4,
    "columnAliases": {
        "Aluminum": ["Aluminium"],
        "Natural gas, US": ["Natural gas, U.S."],
    },
}


def make_workbook_bytes(missing_column: str | None = None) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Monthly Prices"
    worksheet.cell(1, 1, "World Bank Commodity Price Data (The Pink Sheet)")
    worksheet.cell(4, 1, "Updated on July 02, 2026")

    headers = [
        "Crude oil, Brent",
        "Natural gas, US",
        "Aluminum",
        "Iron ore, cfr spot",
        "Copper",
        "Nickel",
        "Zinc",
    ]
    units = ["($/bbl)", "($/mmbtu)", "($/mt)", "($/dmtu)", "($/mt)", "($/mt)", "($/mt)"]
    for column, (header, unit) in enumerate(zip(headers, units), start=2):
        if header == missing_column:
            header = "Missing column"
        worksheet.cell(5, column, header)
        worksheet.cell(6, column, unit)

    year = 2021
    month = 1
    for row_number in range(7, 7 + 66):
        worksheet.cell(row_number, 1, f"{year:04d}M{month:02d}")
        values = [80.5, 3.15, 2500.0, 100.8, 9000.0, 17000.0, 3000.0]
        for column, value in enumerate(values, start=2):
            worksheet.cell(row_number, column, value + row_number)
        month += 1
        if month == 13:
            month = 1
            year += 1

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class WorldBankParserTests(unittest.TestCase):
    def test_resolve_columns_with_aliases(self) -> None:
        headers = [None, "Aluminium", "Natural gas, U.S."]
        materials = [MATERIALS[2], MATERIALS[6]]
        resolved = resolve_columns(headers, materials, CONFIG)
        self.assertEqual(resolved["aluminium"], 2)
        self.assertEqual(resolved["natural_gas"], 3)

    def test_parse_workbook(self) -> None:
        payload = parse_workbook(
            raw=make_workbook_bytes(),
            download_metadata={
                "requestedUrl": "https://thedocs.worldbank.org/test.xlsx",
                "finalUrl": "https://thedocs.worldbank.org/test.xlsx",
                "httpStatus": 200,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentLengthHeader": None,
                "etag": None,
                "lastModified": None,
                "fileSizeBytes": 100000,
                "sha256": "abc",
            },
            materials=MATERIALS,
            config=CONFIG,
            generated_at="2026-07-29T08:00:00Z",
        )
        self.assertTrue(payload["isRealData"])
        self.assertEqual(payload["dataset"]["workbookUpdatedOn"], "2026-07-02")
        self.assertEqual(payload["dataset"]["latestPeriod"], "2026-06")
        self.assertEqual(payload["series"]["iron_ore"]["sourceUnit"], "($/dmtu)")
        self.assertEqual(payload["series"]["copper"]["observationCount"], 66)

    def test_missing_required_column_fails(self) -> None:
        with self.assertRaisesRegex(RuntimeError, "缺少必要欄位"):
            parse_workbook(
                raw=make_workbook_bytes(missing_column="Zinc"),
                download_metadata={"sha256": "abc", "finalUrl": "https://example.invalid"},
                materials=MATERIALS,
                config=CONFIG,
                generated_at="2026-07-29T08:00:00Z",
            )


if __name__ == "__main__":
    unittest.main()
